import datetime
import os

# --- Robust Zero-Dependency Environment Variable Loader ---
def _load_env_manually():
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    val_str = val.strip().strip("'\"")
                    os.environ[key.strip()] = val_str

_load_env_manually()

# pyrefly: ignore [missing-import]
from google.adk import Agent
# pyrefly: ignore [missing-import]
from google.adk.tools.tool_context import ToolContext

# --- Configuration ---
# Google Sheet URL for grounding
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1479hgGqT2Z3E4M-8BvUKj6OOXJMed7P5RcTFsUDiVhE/edit?gid=786787230#gid=786787230"

# Used only if the user's question doesn't mention a region
DEFAULT_REGION = "us-central1"

# --- Tools ---
def pull_payment_sheet_info(tool_context: ToolContext) -> str:
    """Pulls information regarding payment models of Google Cloud (on demand, pt, priority, flex, batch)
    from the Google Sheet, across ALL tabs in the workbook (not just one tab). Automatically records
    the current time as the latestcheck timestamp.

    Args:
        tool_context: The automatically injected ToolContext containing session state.

    Returns:
        A string containing the payment model information from every tab in the sheet, each
        clearly labeled with its tab name.
    """
    import urllib.request
    import re
    import io
    from openpyxl import load_workbook

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tool_context.state['latestcheck'] = now_str


  

    # Parse the spreadsheet ID
    sheet_id_match = re.search(r"/d/([a-zA-Z0-9-_]+)", GOOGLE_SHEET_URL)
    if not sheet_id_match:
        raise ValueError("Could not extract Sheet ID from GOOGLE_SHEET_URL")
    sheet_id = sheet_id_match.group(1)

    # Export the ENTIRE workbook as .xlsx so every tab (on-demand, PT, priority,
    # flex, batch, etc.) comes back in one fetch, instead of exporting a single
    # gid's tab as CSV.
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    req = urllib.request.Request(
        export_url,
        headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    )
    with urllib.request.urlopen(req, timeout=20) as response:
        xlsx_bytes = response.read()

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                rows.append(",".join("" if c is None else str(c) for c in row))
        if rows:
            sections.append(f"--- TAB: {sheet_name} ---\n" + "\n".join(rows))

    if not sections:
        raise ValueError("Workbook fetched but no non-empty tabs found")

    # tool_context.state["response" ] = chr(10).join(sections)

    return f"""
    Latest Check Time: {now_str}
    Spreadsheet Source: {GOOGLE_SHEET_URL}

    {chr(10).join(sections)}
    """


# --- Root Agent ---
# Standard ADK Agent, discovered directly by the ADK CLI/Web UI — no custom
# Workflow/Event graph needed. ADK's own runner handles the event loop.
root_agent = Agent(
    name="find_info",
    model="gemini-2.5-flash",
    instruction=f"""always call the `pull_payment_sheet_info` 
    answer all questions after that
""",
    tools=[pull_payment_sheet_info],
    description="Retrieves information about Google Cloud payment models and pricing from every tab of the Google Sheet and answers questions.",
)